"""
Reads in an Excel document to add blank lines.
The user can run the program in command line with arguments or input
the information at runtime.
Things to input:
-Filename of excel file to manipulate
-Start row for blank lines
-Count of blank rows to insert

The changes made by the program get saved in a new excel file.
It is saved as updated_{filename}.xlsx
"""

import sys
from typing import Tuple
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

def is_positive_number(input_string: str) -> bool:
    """Checks if input is valid to use for the excel file"""
    return input_string.isdigit() and int(input_string) >= 1


def enter_positive_number(user_message) -> int:
    """
    Prompts the user until positive number was entered and returns then
    the number
    """
    while True:
        input_string = input(user_message)
        if is_positive_number(input_string):
            return int(input_string)


def user_input() -> Tuple[int, int, str]:
    """
    First looks for data on the command line. If command line input is
    not valid the user is prompted to put in the values manually.
    Returns start_row, count_of_blank_rows and filename
    """
    if len(sys.argv) == 4:
        start_blank_row_input: str = sys.argv[1]
        blank_rows_input: str = sys.argv[2]

        if (is_positive_number(start_blank_row_input) and
                is_positive_number(blank_rows_input)):
            return (int(start_blank_row_input),
                    int(blank_rows_input), sys.argv[3])

    start_blank_row = enter_positive_number(
        "Enter start of blank rows:\n")
    blank_rows = enter_positive_number(
        "Enter how many blank rows to insert:\n")
    filename: str = input("Enter filename:\n")
    return start_blank_row, blank_rows, filename


def save_workbook_excel_file(workbook, filename):
    """Tris to save created data to excel file"""
    try:
        workbook.save(filename)
    except PermissionError:
        print("Error: No permission to save file.")


def copy_cells(
        source_sheet, target_sheet,
        row_start: int, row_end: int, row_offset: int,
        column_start: int, column_end: int, column_offset: int):
    """
    Copies cells from one sheet to the other including its style.
    It is possible to enter blank rows or columns into the new sheet.
    """
    for row in range(row_start, row_end):
        for column in range(column_start, column_end):
            source = source_sheet.cell(row=row, column=column)
            target = target_sheet.cell(
                row=row + row_offset, column=column + column_offset)
            target.value = copy(source.value)
            if source.has_style:
                target.font = copy(source.font)
                target.border = copy(source.border)
                target.fill = copy(source.fill)
                target.number_format = copy(source.number_format)
                target.protection = copy(source.protection)
                target.alignment = copy(source.alignment)


def copy_row_dimensions(
        source_sheet, target_sheet,
        start_blank_row: int, blank_rows: int):
    """
    Copies the dimension of rows from one sheet to the other.
    It is possible to copy with an offset to "insert" new rows
    into the new sheet
    """
    for row_number, row_dim in source_sheet.row_dimensions.items():
        if row_number >= start_blank_row:
            row_number = row_number + blank_rows
        target_sheet.row_dimensions[row_number] = copy(row_dim)


def copy_column_dimensions(
        source_sheet, target_sheet,
        start_blank_column: int, blank_columns: int):
    """
    Copies the dimension of columns from one sheet to the other.
    It is possible to copy with an offset to "insert" new columns
    into the new sheet
    """
    for column_letter, column_dim in source_sheet.column_dimensions.items():
        column_index = column_index_from_string(column_letter)
        if column_index >= start_blank_column:
            column_index = column_index + blank_columns
        column_letter = get_column_letter(column_index)
        target_sheet.column_dimensions[column_letter] = copy(column_dim)


def blank_row_inserter():
    """Main logic to insert blank rows"""
    start_blank_row: int
    blank_rows: int
    filename: str
    start_blank_row, blank_rows, filename = user_input()

    workbook = openpyxl.load_workbook(filename)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]
    max_row = sheet.max_row
    max_column = sheet.max_column
    workbook.create_sheet(index=0, title='tmp_sheet')
    new_sheet = workbook['tmp_sheet']

    # Copy everything before blank area
    copy_cells(
        source_sheet=sheet, target_sheet=new_sheet,
        row_start=1, row_end=start_blank_row, row_offset=0,
        column_start=1, column_end=max_column, column_offset=0)

    # Copy with row offset
    copy_cells(
        source_sheet=sheet, target_sheet=new_sheet,
        row_start=start_blank_row, row_end=max_row, row_offset=blank_rows,
        column_start=1, column_end=max_column, column_offset=0)

    copy_row_dimensions(
        source_sheet=sheet, target_sheet=new_sheet,
        start_blank_row=start_blank_row, blank_rows=blank_rows)

    copy_column_dimensions(
        source_sheet=sheet, target_sheet=new_sheet,
        start_blank_column=0, blank_columns=0)

    sheet_name = sheet.title
    del workbook[sheet_name]
    new_sheet.title = sheet_name
    save_workbook_excel_file(workbook, 'updated_' + filename)


if __name__ == "__main__":
    blank_row_inserter()
