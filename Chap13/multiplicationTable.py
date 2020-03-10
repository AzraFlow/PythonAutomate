#! Python3
# Code to accept an integer as a command line argument and create an Excel file with a multiplication table of that size (nXn)
import logging
logging.disable(logging.CRITICAL)  # Uncomment to disable debug logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

import sys, openpyxl
from openpyxl.styles import Font

# Get matrix size from the command line
if len(sys.argv) != 2:
    print('Usage: python multiplicationTable.py [matrix_size]')
    sys.exit()
matrixSize = int(sys.argv[1])
logging.debug('matrixSize (%s)' %(matrixSize))

# Create an Excel workbook and sheet
wb = openpyxl.Workbook()  # Create a blank Workbook
wb.sheetnames  # Start with one sheet
sheet = wb.active
sheet.title = 'Multiplication Table'

# Create axis labels for Table in Bold font
fontObj1 = Font(name='Times New Roman', size="14", bold=True)
for colLabel in range(1, matrixSize + 1):
    sheet.cell(row=1, column=colLabel+1).font = fontObj1
    sheet.cell(row=1, column=colLabel+1).value = colLabel
for rowLabel in range(1, matrixSize + 1):
    sheet['A' + str(rowLabel + 1)].font = fontObj1
    sheet['A' + str(rowLabel + 1)] = rowLabel

# Populate sheet with multiplication table data
for rowNum in range(1, matrixSize + 1):
    for columnNum in range(1, matrixSize + 1):
        sheet.cell(row=rowNum+1, column=columnNum+1).value = rowNum * columnNum

# Save Excel workbook
wb.save('multiplicationTable.xlsx')
